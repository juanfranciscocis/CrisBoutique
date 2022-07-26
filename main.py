import os
import subprocess
import tkinter
import sys

from tkinter import *
from tkinter import filedialog as fd
from tkinter import messagebox
from tkinter import ttk
from tkinter.ttk import Combobox

import openpyxl

# ABRO EL ARCHIVO EXCEL FUNCION
from openpyxl.styles import PatternFill

filename = ""


def inputExcel():
    global filename
    filename = fd.askopenfilename()
    print(filename)
    # ACTIVO EL EXCEL
    global wb
    wb = openpyxl.load_workbook(filename.replace(" ", ""))
    # global ws
    # ws = wb.active


# RESTART FUNCTION
def restartProgramAndSave():
    os.execl(sys.executable, sys.executable, *sys.argv)


def addClothesAndBills():
    # CREO VENTANA ADD CLOTHES AND BILLS
    global rootMain
    rootMain = Toplevel(root)
    rootMain.title('CRIS BOUTIQUE BILLS AND CLOTHES INDEX')
    rootMain.config(background='#FDE5E6')
    rootMain.geometry("1280x700")
    
    rootMain.resizable(0, 0)

    img = PhotoImage(file='FONDO-3.png')
    fondo = Label(rootMain, image=img)
    fondo.config(background='#FDE5E6')
    fondo.place(x=0, y=0)

    if filename == "":
        print("error filename")
        r = tkinter.messagebox.askquestion(
            message="ARCHIVO DE EXCEL NO AGREGADO!!! DESEA AGREGAR UN ARCHIVO AHORA?", title="HEY!!!")
        if (r == 'yes'):
            inputExcel()

    # FONT
    fontTuple = ("Strawberry Cupcakes", 16, "bold")
    fontTupleEntry = ("Gotham Medium", 16)
    fontTupleEntry2 = ("Gotham Medium", 20)

    # DISPLAY Y OBTENCION DE USUARIO DEL CODIGO DE HOJA EXCEL
    # literalCodigoLabel = Text(rootMain, background = '#F9D4D4', height = 1, width = 35)
    # literalCodigoLabel.config(highlightbackground= '#F9D4D4')
    # literalCodigoLabel.configure(font = fontTuple)
    # literalCodigoLabel.insert(END, 'INGRESE EL  LITERAL  DEL  CODIGO')
    # literalCodigoLabel.grid(column = 0, row = 0)
    # literalCodigoLabel.config(state = DISABLED)

    literalCodigoEntry = Text(rootMain, height=1, width=2, background="#FDE5E6")
    literalCodigoEntry.configure(font=fontTupleEntry)
    literalCodigoEntry.tag_config("center", justify='center')
    literalCodigoEntry.config(highlightbackground="#FDE5E6")
    literalCodigoEntry.place(x=440, y=33) #540-46
    usr_codigoLiteral = literalCodigoEntry.get("1.0", "end-1c")
    literalCodigoEntry.tag_add("center", "1.0", "end")

    # DISPLAY Y OBTENCION DE USUARIO DEL CODIGO DE INICIO DE FACTURA
    # numberCodigoLabel = Text(rootMain, background = '#F9D4D4', height = 1, width = 33)
    # numberCodigoLabel.configure(font = fontTuple)
    # numberCodigoLabel.config(highlightbackground= '#F9D4D4')
    # numberCodigoLabel.insert(END,' CODIGO DE INICIO DE  FACTURA ')
    # numberCodigoLabel.place(x=480, y = 2)
    # numberCodigoLabel.config(state = DISABLED)

    numberCodigoEntry = Text(rootMain, height=1, width=3, background="#FDE5E6")
    numberCodigoEntry.configure(font=fontTupleEntry)
    numberCodigoEntry.config(highlightbackground="#FDE5E6")
    numberCodigoEntry.place(x=940, y=33)
    numberCodigoEntryVar = numberCodigoEntry.get("1.0", "end-1c")

    # DISPLAY Y OBTENCION DE USUARIO DE LA MARCA
    # brandLabel = Text(rootMain, background = '#F9D4D4', height = 1, width = 33)
    # brandLabel.configure(font = fontTuple)
    # brandLabel.config(highlightbackground= '#F9D4D4')
    # brandLabel.insert(END,'MARCA')
    # brandLabel.place(x=1460, y =2 )
    # brandLabel.config(state = DISABLED)

    brandEntry = Text(rootMain, height=1, width=20, background="#FDE5E6")
    brandEntry.configure(font=fontTupleEntry2)
    brandEntry.config(highlightbackground="#FDE5E6")
    brandEntry.place(x=110, y=490)

    # DISPLAY MENU AGREGAR FACTURAS
    # priceTitleMenu = ttk.Label(rootMain, text = "AGREGAR UNA FACTURA AL EXCEL", font = ("Arial", 27))
    # priceTitleMenu.grid(column = 3, row =2 )

    # DISPLAY Y OBTENCION DE PRECIO DE FACTURA
    # priceLabel = ttk.Label(rootMain, text = "PRECIO DE FACTURA $")
    # priceLabel.grid(column = 3 , row = 3)
    priceEntry = ttk.Entry(rootMain, width=3, background="#FDE5E6")
    priceEntry.configure(font=fontTupleEntry2)
    priceEntry.place(x=145, y=570)

    # DISPLAY RECUADRO DE IMPRESION DE DATA
    printData = Text(rootMain, height=10, width= 55 , font=fontTupleEntry, wrap=WORD, background="#FDE5E6")
    printData.place(x=600, y=445)
    printData.config(highlightbackground="#FDE5E6")
    printData.insert(END, "\n")
    printData.insert(END,
                     "----------------------------------------------------------------------------------------------------------------------------")
    printData.insert(END, "\n")
    printData.insert(END, "\n")
    printData.insert(END, "\n")
    printData.insert(END, "\n")
    printData.insert(END, "LOS DATOS QUE NO APARECEN EN ESTE RECUADRO NO SON AGREGADOS AL EXCEL NORMALMENTE DEBIDO A "
                          "ERRORES DESCONOCIDOS O ERRORES PRODUCIDOS POR EL USUARIO!")
    printData.insert(END, "\n")
    printData.insert(END, "\n")
    printData.insert(END, "\n")
    printData.insert(END, "\n")
    printData.insert(END,
                     "----------------------------------------------------------------------------------------------------------------------------")
    printData.config(state=DISABLED)  # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION

    # BUSQUEDA DEL CODIGO EN EXCEL
    def wordfinder(searchString):
        while (True):
            buscarColumn = 0
            buscarFila = 0
            count = 0

            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    if searchString == ws.cell(i, j).value:
                        print("SE ENCONTRO EL CODIGO EN LA HOJA EXCEL")
                        print("/////////////////////////////////////////")
                        # print(ws.cell(i,j))
                        buscarColumn = i
                        buscarFila = j
                        count += 1
                        # print("Columna ", buscarColumn)
                        # print("Fila", buscarFila)

            if buscarFila == 0 and buscarFila == 0:
                tkinter.messagebox.showinfo(message="NO SE ENCUENTRA EL CODIGO", title="HEY!!!")
                print("NO SE ENCONTRO EL CODIGO EN LA HOJA DE EXCEL")
                buscarColumn = 0
                buscarFila = 0
                print("/////////////////////////////////////////")

                break

            return buscarFila, buscarColumn, count

    # CAMBIO DE NUMERO DE COLUMNA A LETRA DE COLUMNA
    def cambioABC(cambio):
        if cambio == 1:
            cambioNumaLetra = "A"
        elif cambio == 2:
            cambioNumaLetra = "B"
        elif cambio == 3:
            cambioNumaLetra = "C"
        elif cambio == 4:
            cambioNumaLetra = "D"
        elif cambio == 5:
            cambioNumaLetra = "E"
        elif cambio == 6:
            cambioNumaLetra = "F"
        elif cambio == 7:
            cambioNumaLetra = "G"
        elif cambio == 8:
            cambioNumaLetra = "H"
        elif cambio == 9:
            cambioNumaLetra = "I"
        elif cambio == 10:
            cambioNumaLetra = "J"
        elif cambio == 11:
            cambioNumaLetra = "K"
        elif cambio == 12:
            cambioNumaLetra = "L"
        elif cambio == 13:
            cambioNumaLetra = "M"
        elif cambio == 14:
            cambioNumaLetra = "N"
        elif cambio == 15:
            cambioNumaLetra = "O"
        elif cambio == 16:
            cambioNumaLetra = "P"
        elif cambio == 17:
            cambioNumaLetra = "Q"
        elif cambio == 18:
            cambioNumaLetra = "R"
        elif cambio == 19:
            cambioNumaLetra = "S"
        elif cambio == 20:
            cambioNumaLetra = "T"
        elif cambio == 21:
            cambioNumaLetra = "U"
        elif cambio == 22:
            cambioNumaLetra = "V"
        elif cambio == 23:
            cambioNumaLetra = "W"
        elif cambio == 24:
            cambioNumaLetra = "X"
        elif cambio == 25:
            cambioNumaLetra = "Y"
        elif cambio == 26:
            cambioNumaLetra = "Z"
        else:
            cambioNumaLetra = "A"

        return cambioNumaLetra

    # RETORNO DE CELDA CON COORDENADAS
    def val(x, y):
        return ws.cell(row=x, column=y).value

    # CODIGO DE LA PLANTILLA EXCEL COMPLETO ALFANUMERICO
    def addProduct():
        while (True):
            global ws
            ws = wb.active

            # PERMITO LA ESCRITURA AL WIDGET DE IMPRESION
            printData.config(state=NORMAL)

            # MEJORO EL MANEJO DE LA OBTENCION DE DATA DEL WIDGET
            numberCodigoEntryVar = numberCodigoEntry.get("1.0", "end-1c")

            codigoFinal = literalCodigoEntry.get("1.0", "end-1c").upper() + "." + str(
                numberCodigoEntryVar)  # CODIGO DE PLANTILLA EXCEL
            literalCodigoEntry.tag_add("tag_name", "1.0", "end")
            print("COMIENZO DESDE EL CODIGO: ", codigoFinal)

            # CHEQUEO QUE LOS WIDGETS DE MARCA Y PRECIO TENGAN INPUT
            if (brandEntry.get("1.0", "end-1c") == ""):
                tkinter.messagebox.showinfo(message="FALTA AGREGAR LA MARCA, EL ARTICULO NO HA SIDO AGREGADO!",
                                            title="HEY!!!")
                # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                printData.config(state=DISABLED)
                break
            if (priceEntry.get() == ""):
                tkinter.messagebox.showinfo(message="FALTA AGREGAR EL PRECIO,  EL ARTICULO NO HA SIDO AGREGADO!",
                                            title="HEY!!!")
                # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                printData.config(state=DISABLED)
                break

            cordenada_excel = wordfinder(codigoFinal)  # CAMBIO A COORDENADAS CON FUNCION WORDFINDER

            cordenada_excel_codigo = cambioABC(cordenada_excel[0]) + str(cordenada_excel[1])
            cordenada_excel_marca = cambioABC(cordenada_excel[0] + 1) + str(cordenada_excel[1])
            cordenada_excel_precioFactura = cambioABC(cordenada_excel[0] + 3) + str(cordenada_excel[1])
            cordenada_excel_tax = cambioABC(cordenada_excel[0] + 4) + str(cordenada_excel[1])
            cordenada_excel_pu = cambioABC(cordenada_excel[0] + 5) + str(cordenada_excel[1])
            cordenada_excel_cincoPorciento = cambioABC(cordenada_excel[0] + 6) + str(cordenada_excel[1])
            cordenada_excel_costo = cambioABC(cordenada_excel[0] + 7) + str(cordenada_excel[1])

            if ws[cordenada_excel_marca].value != None:
                r = tkinter.messagebox.askquestion(
                    message="YA EXISTE UNA FACTURA INGRESADA CON ESTE CODIGO. DESEA SOBREESCRIBIR ?", title="HEY!!!")
                if (r == 'no'):
                    mostrarYAbrirArchivo = "DUPLICADO ENCONTRADO EN LA CELDA: " + str(cordenada_excel_codigo)
                    tkinter.messagebox.showinfo(message=mostrarYAbrirArchivo, title="HEY!!!")
                    subprocess.run(['open', filename], check=True)
                    # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                    printData.config(state=DISABLED)
                    break

                print("SI")
                print()
                print("/////////////////////////////////////////")
                print("YA EXISTE UNA MARCA CON ESTE CODIGO, CHEQUEE EL EXCEL!!!")
                print("CODIGO DE VENTA: ", val(cordenada_excel[1], cordenada_excel[0]))
                print("/////////////////////////////////////////")
                print()
                print()

            total = 0
            tax = 0.065
            precioUnitario = 0
            cincoPorciento = 0.05

            ws[cordenada_excel_marca] = brandEntry.get("1.0", "end-1c").upper()
            print(ws[cordenada_excel_marca])
            ws[cordenada_excel_precioFactura] = float(priceEntry.get())
            total += float(priceEntry.get())
            ws[cordenada_excel_tax] = tax * float(priceEntry.get())  # agrego tax

            # agrego PU

            precioUnitario = (tax * float(priceEntry.get())) + float(priceEntry.get())
            ws[cordenada_excel_pu] = precioUnitario

            # agrego 5%
            ws[cordenada_excel_cincoPorciento] = precioUnitario * cincoPorciento

            # agrego costo
            ws[cordenada_excel_costo] = (precioUnitario * cincoPorciento) + precioUnitario

            # IMPRIMO DATA EN EL WIDGET DE IMPRESION
            printData.insert(END,
                             "----------------------------------------------------------"
                             "------------------------------------------------------------------")
            printData1 = "SE AGREGO CORRECTAMENTE A LA MARCA: " + str(
                val(cordenada_excel[1], cordenada_excel[0] + 1).replace(" ", ""))
            printData.insert(END, "\n")
            printData.insert(END, printData1)
            printData.insert(END, "\n")
            printData2 = "EL PRECIO DEL ARTICULO ES: " + str(val(cordenada_excel[1], cordenada_excel[0] + 3)) + "$"
            printData.insert(END, printData2)
            printData.insert(END, "\n")
            printData3 = "CODIGO DE VENTA: " + str(val(cordenada_excel[1], cordenada_excel[0]))
            printData.insert(END, printData3)
            printData.insert(END, "\n")
            printData.insert(END,
                             "------------------------------------------------------------"
                             "----------------------------------------------------------------")
            printData.yview(END)

            # ELIMINO TEXTO DEl WIDGET numberCodigoEntry y priceLabelEntry
            numberCodigoEntry.delete("1.0", "end")
            priceEntry.delete(0, END)

            # SUMO 1 A EL WIDGET numberCodigoEntry
            numberCodigoEntryADD = int(numberCodigoEntryVar) + 1
            numberCodigoEntry.insert(END, str(numberCodigoEntryADD))

            # GUARDO EL ARCHIVO CON EL PRODUCTO NUEVO AGREGADO
            wb.save(filename.replace(" ", ""))
            ws = wb.close

            # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
            printData.config(state=DISABLED)

            break

        # BOTON DE AGREGAR A EXCEL

    def showProduct():
        global ws
        ws = wb.active
        global numberCodigoEntryVar
        # PERMITO LA ESCRITURA AL WIDGET DE IMPRESION
        printData.config(state=NORMAL)

        # MEJORO EL MANEJO DE LA OBTENCION DE DATA DEL WIDGET
        numberCodigoEntryVar = numberCodigoEntry.get("1.0", "end-1c")

        codigoFinal = literalCodigoEntry.get("1.0", "end-1c").upper() + "." + str(
            numberCodigoEntryVar)  # CODIGO DE PLANTILLA EXCEL
        literalCodigoEntry.tag_add("tag_name", "1.0", "end")
        print("COMIENZO DESDE EL CODIGO: ", codigoFinal)
        # IMPRESION DE LA CELDA ENCONTRADA
        cordenada_Excel = wordfinder(codigoFinal)  # CAMBIO A COORDENADAS CON FUNCION WORDFINDER

        print()
        print()
        printData.insert(END, "\n")
        printData.insert(END, "\n")
        printData.insert(END, "---------------------------------------------")
        printData.insert(END, "\n")
        printData3 = "MARCA DEL ARTICULO---PRECIO DE LA FACTURA"
        printData.insert(END, printData3)
        printData.insert(END, "\n")
        printData4 = str(val(cordenada_Excel[1], cordenada_Excel[0] + 1)) + "--------" + str(
            val(cordenada_Excel[1], cordenada_Excel[0] + 3)) + "$"
        printData.insert(END, printData4)
        printData.insert(END, "\n")
        printData.insert(END,
                         "-------------------------------------------------------------------------------------"
                         "---------------------------------------")
        printData.yview(END)
        printData.config(state=DISABLED)

        ws = wb.close

    # boton Search
    showProductButton = Button(rootMain, text=" SEARCH ", command=showProduct)
    showProductButton.config(highlightbackground="#FFB1BE")
    showProductButton.place(x=1000, y=33)

    # BOTON DE AGREGAR A EXCEL
    add = Button(rootMain, text=" ADD PRODUCT ", command=addProduct)
    add.config(highlightbackground="#FDE5E6")
    add.place(x=390, y=580) #400,580
    # -----------------------------------------------------------------------------------------------------------------

    # EXIT PROGRAM AND SAVE FUNCTION
    def exitProgramAndSave():
        print("GUARDANDO ARCHIVO EXCEL Y CERRANDO APP")
        wb.save(filename.replace(" ", ""))
        sys.exit()

    # EXIT BUTTON
    exitProgram = Button(rootMain, text=" CLOSE AND SAVE ", command=exitProgramAndSave)
    exitProgram.config(highlightbackground="#FFB1BE")
    exitProgram.place(x=485, y=640)

    # RESTART BUTTON
    restartProgram = Button(rootMain, text="OPEN ANOTHER EXCEL SHEET OR RESTART APP", command=restartProgramAndSave)
    restartProgram.config(highlightbackground="#FFB1BE")
    restartProgram.place(x=810, y=640)

    # OPEN EXCEL FILE FUNCTION
    def openFile():
        subprocess.run(['open', filename], check=True)

    # OPEN EXCEL FILE BUTTON
    openFileButton = Button(rootMain, text="OPEN CURRENT EXCEL SHEET", command=openFile)
    openFileButton.config(highlightbackground="#FFB1BE")
    openFileButton.place(x=200, y=640)

    # -----------------------------------------------------------------------------------------------------------------
    # PRICE ENTRY 2 FOR ADDING PERSONAL PRODUCTS
    priceEntry2 = ttk.Entry(rootMain, width=3, background="#FDE5E6")
    priceEntry2.configure(font=fontTupleEntry2)
    priceEntry2.place(x=145, y=148)

    # COLOR PICKER -----------------------------------------------------------------------------------------------------
    colorPickedOriginal = ""

    # FUNCTION RED BUTTON
    def redColorPicked():
        colorPicked = "ROJA"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked

        print(colorPicked)

    # COLOR RED BUTTON
    redButton = Button(rootMain, text='   ', command=redColorPicked)
    redButton.place(x=100, y=270)
    redButton.config(highlightbackground="#FFB1BE")

    # FUNCTION YELLOW BUTTON
    def yellowColorPicked():
        colorPicked = "AMARILLO"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR YELLOW BUTTON
    yellowButton = Button(rootMain, text='   ', command=yellowColorPicked)
    yellowButton.place(x=160, y=270)
    yellowButton.config(highlightbackground="#FFB1BE")

    # FUNCTION BLUE BUTTON
    def blueColorPicked():
        colorPicked = "AZUL"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR BLUE BUTTON
    blueButton = Button(rootMain, text='   ', command=blueColorPicked)
    blueButton.place(x=230, y=270)
    blueButton.config(highlightbackground="#FFB1BE")

    # FUNCTION PINK BUTTON
    def pinkColorPicked():
        colorPicked = "ROSADA"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR PINK BUTTON
    pinkButton = Button(rootMain, text='   ', command=pinkColorPicked)
    pinkButton.place(x=280, y=270)
    pinkButton.config(highlightbackground="#FFB1BE")

    # FUNCTION TOMATE BUTTON
    def orangeColorPicked():
        colorPicked = "TOMATE"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR TOMATE BUTTON
    orangeButton = Button(rootMain, text='   ', command=orangeColorPicked)
    orangeButton.place(x=340, y=270)
    orangeButton.config(highlightbackground="#FFB1BE")

    # FUNCTION VERDE BUTTON
    def greenColorPicked():
        colorPicked = "VERDE"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR VERDE BUTTON
    greenButton = Button(rootMain, text='   ', command=greenColorPicked)
    greenButton.place(x=395, y=270)
    greenButton.config(highlightbackground="#FFB1BE")

    # FUNCTION MORADO BUTTON
    def purpleColorPicked():
        colorPicked = "MORADO"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR MORADO BUTTON
    purpleButton = Button(rootMain, text='   ', command=purpleColorPicked)
    purpleButton.place(x=440, y=270)
    purpleButton.config(highlightbackground="#FFB1BE")

    # FUNCTION CAFE BUTTON
    def brownColorPicked():
        colorPicked = "CAFE"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR MORADO BUTTON
    brownButton = Button(rootMain, text='   ', command=brownColorPicked)
    brownButton.place(x=500, y=270)
    brownButton.config(highlightbackground="#FFB1BE")

    # FUNCTION NEGRO BUTTON
    def blackColorPicked():
        colorPicked = "NEGRO"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR NEGRO BUTTON
    blackButton = Button(rootMain, text='   ', command=blackColorPicked)
    blackButton.place(x=550, y=270)
    blackButton.config(highlightbackground="#FFB1BE")

    # FUNCTION BLANCO BUTTON
    def whiteColorPicked():
        colorPicked = "BLANCO"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR BLANCO BUTTON
    whiteButton = Button(rootMain, text='   ', command=whiteColorPicked)
    whiteButton.place(x=600, y=270)
    whiteButton.config(highlightbackground="#FFB1BE")

    # FUNCTION BLANCO Y NEGROBUTTON
    def whiteAndBlackColorPicked():
        colorPicked = "BLANCO Y NEGRO"
        global colorPickedOriginal
        colorPickedOriginal = colorPicked
        print(colorPicked)

    # COLOR BLANCO BUTTON
    whiteAndBlackButton = Button(rootMain, text='   ', command=whiteAndBlackColorPicked)
    whiteAndBlackButton.place(x=650, y=270)
    whiteAndBlackButton.config(highlightbackground="#FFB1BE")

    # SHAPE PICKER ----------------------------------------------------------------------------------------------------

    shapePickedOriginal = ""

    # FUNCTION BOLITAS SHAPE
    def ballsTypePicked():
        shapePicked = "BOLITAS"
        global shapePickedOriginal
        shapePickedOriginal = shapePicked
        print(shapePicked)

    # BOTON BOLITAS SHAPE
    bolitasShapePicker = Button(rootMain, text='   ', command=ballsTypePicked)
    bolitasShapePicker.config(highlightbackground="#FFB1BE")
    bolitasShapePicker.place(x=94, y=380)

    # FUNCTION RAYAS SHAPE
    def stripesTypePicked():
        shapePicked = "RAYAS"
        global shapePickedOriginal
        shapePickedOriginal = shapePicked
        print(shapePicked)

    # BOTON RAYAS SHAPE
    rayasShapePicker = Button(rootMain, text='   ', command=stripesTypePicked)
    rayasShapePicker.config(highlightbackground="#FFB1BE")
    rayasShapePicker.place(x=163, y=380)

    # FUNCTION FLORES SHAPE
    def flowersTypePicked():
        shapePicked = "FLORES"
        global shapePickedOriginal
        shapePickedOriginal = shapePicked
        print(shapePicked)

    # BOTON FLORES SHAPE
    florShapePicker = Button(rootMain, text='   ', command=flowersTypePicked)
    florShapePicker.config(highlightbackground="#FFB1BE")
    florShapePicker.place(x=230, y=380)

    # FUNCTION JASPEADO SHAPE
    def jaspeadoTypePicked():
        shapePicked = "JASPEADO"
        global shapePickedOriginal
        shapePickedOriginal = shapePicked
        print(shapePicked)

    # BOTON JASPEADO SHAPE
    jaspeadoShapePicker = Button(rootMain, text='   ', command=jaspeadoTypePicked)
    jaspeadoShapePicker.config(highlightbackground="#FFB1BE")
    jaspeadoShapePicker.place(x=280, y=380)

    # FUNCTION MANDALA SHAPE
    def mandalaTypePicked():
        shapePicked = "MANDALA"
        global shapePickedOriginal
        shapePickedOriginal = shapePicked
        print(shapePicked)

    # BOTON MANDALA SHAPE
    mandalaShapePicker = Button(rootMain, text='   ', command=mandalaTypePicked)
    mandalaShapePicker.config(highlightbackground="#FFB1BE")
    mandalaShapePicker.place(x=330, y=380)

    # FUNCTION NONE SHAPE
    def noneTypePicked():
        shapePicked = ""
        global shapePickedOriginal
        shapePickedOriginal = shapePicked
        print(shapePicked)

    # BOTON NONE SHAPE
    noneShapePicker = Button(rootMain, text='   ', command=noneTypePicked)
    noneShapePicker.config(highlightbackground="#FFB1BE")
    noneShapePicker.place(x=390, y=380)

    # TYPE PICKER ---------------------------------------------------------------------------------------------------------

    typePickedOriginal = ""

    # FUNCTION BLUSA TYPE
    def blouseTypePicked():
        typePicked = "BLUSA"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)
        print(typePickedOriginal + "ESTE")

    # BOTON BLUSA TYPE
    blusaTypePicker = Button(rootMain, text=' BLUSA ', command=blouseTypePicked)
    blusaTypePicker.config(highlightbackground="#FDE5E6")
    blusaTypePicker.place(x=380, y=150)

    # FUNCTION CARDIGAN TYPE
    def cardiganTypePicked():
        typePicked = "CARDIGAN"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON CARDIGAN TYPE
    cardiganTypePicker = Button(rootMain, text=' CARDIGAN ', command=cardiganTypePicked)
    cardiganTypePicker.config(highlightbackground="#FDE5E6")
    cardiganTypePicker.place(x=470-15, y=150)

    # FUNCTION PANTALON TYPE
    def pantalonTypePicked():
        typePicked = "PANTALON"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON PANTALON TYPE
    pantalonTypePicker = Button(rootMain, text=' PANTALON ', command=pantalonTypePicked)
    pantalonTypePicker.config(highlightbackground="#FDE5E6")
    pantalonTypePicker.place(x=660-80-20-15, y=150)

    # FUNCTION CAMISA TYPE
    def camisaTypePicked():
        typePicked = "CAMISA"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON CAMISA TYPE
    camisaTypePicker = Button(rootMain, text=' CAMISA ', command=camisaTypePicked)
    camisaTypePicker.config(highlightbackground="#FDE5E6")
    camisaTypePicker.place(x=760-80-20-15, y=150)

    # FUNCTION SOSTEN TYPE
    def sostenTypePicked():
        typePicked = "SOSTEN"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON SOSTEN TYPE
    sostenTypePicker = Button(rootMain, text=' SOSTEN ', command=sostenTypePicked)
    sostenTypePicker.config(highlightbackground="#FDE5E6")
    sostenTypePicker.place(x=840-80-20-15, y=150)

    # FUNCTION CALZON TYPE
    def calzonTypePicked():
        typePicked = "CALZON"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON CALZON TYPE
    calzonTypePicker = Button(rootMain, text=' CALZON ', command=calzonTypePicked)
    calzonTypePicker.config(highlightbackground="#FDE5E6")
    calzonTypePicker.place(x=920-80-20-15, y=150)

    # FUNCTION MEDIAS TYPE
    def mediasTypePicked():
        typePicked = "MEDIAS"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON MEDIAS TYPE
    mediasTypePicker = Button(rootMain, text=' MEDIAS ', command=mediasTypePicked)
    mediasTypePicker.config(highlightbackground="#FDE5E6")
    mediasTypePicker.place(x=1000-80-20-15, y=150)

    # FUNCTION POLO TYPE
    def poloTypePicked():
        typePicked = "POLO"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON POLO TYPE
    poloTypePicker = Button(rootMain, text=' POLO ', command=poloTypePicked)
    poloTypePicker.config(highlightbackground="#FDE5E6")
    poloTypePicker.place(x=1070-80-20-15, y=150)

    # FUNCTION BOXER TYPE
    def boxerTypePicked():
        typePicked = "BOXER"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON BOXER TYPE
    boxerTypePicker = Button(rootMain, text=' BOXER ', command=boxerTypePicked)
    boxerTypePicker.config(highlightbackground="#FDE5E6")
    boxerTypePicker.place(x=1130-80-20-15, y=150)

    # FUNCTION ZAPATOS TYPE
    def zapatosTypePicked():
        typePicked = "CAMISETA"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON ZAPATOS TYPE
    zapatosTypePicker = Button(rootMain, text=' CAMISETA ', command=zapatosTypePicked)
    zapatosTypePicker.config(highlightbackground="#FDE5E6")
    zapatosTypePicker.place(x=1200-80-20-15, y=150)

    # FUNCTION SACO TYPE
    def sacoTypePicked():
        typePicked = "SACO"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON SACO TYPE
    sacoTypePicker = Button(rootMain, text=' SACO ', command=sacoTypePicked)
    sacoTypePicker.config(highlightbackground="#FDE5E6")
    sacoTypePicker.place(x=920, y=180)

    # FUNCTION CHOMPA TYPE
    def chompaTypePicked():
        typePicked = "CHOMPA"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON CHOMPA TYPE
    chompaTypePicker = Button(rootMain, text=' CHOMPA ', command=chompaTypePicked)
    chompaTypePicker.config(highlightbackground="#FDE5E6")
    chompaTypePicker.place(x=920+120, y=180)

    # FUNCTION CORREA TYPE
    def correaTypePicked():
        typePicked = "CORREA"
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # BOTON CORREA TYPE
    correaTypePicker = Button(rootMain, text=' CORREA ', command=correaTypePicked)
    correaTypePicker.config(highlightbackground="#FDE5E6")
    correaTypePicker.place(x=920+120+120, y=180)

    # OTRO TYPE ENTRY
    anotherTypeEntry = Text(rootMain, height=1, width=29, background="#FDE5E6")
    anotherTypeEntry.configure(font=fontTupleEntry2)
    anotherTypeEntry.config(highlightbackground="#FDE5E6")
    anotherTypeEntry.place(x=520, y=115)

    # OTRO TYPE FUCNTION
    def other():
        typePicked = anotherTypeEntry.get("1.0", "end-1c").upper()
        global typePickedOriginal
        typePickedOriginal = typePicked
        print(typePicked)

    # OTRO TYPE FUCNTION BUTTON
    anotherTypeButton = Button(rootMain, text=" OTHER ", command=other)
    anotherTypeButton.place(x=930, y=115)
    anotherTypeButton.config(highlightbackground="#FDE5E6")

    # VARIABLE FINAL (COLOR, SHAPE Y TYPE)
    finalItem = ""

    def finalItems():
        global finalItem
        global typePickedOriginal
        global shapePickedOriginal
        global colorPickedOriginal
        finalItem = str(typePickedOriginal) + " " + str(shapePickedOriginal) + " " + str(colorPickedOriginal)
        print(finalItem)

    def addItem():
        while (True):
            global typePickedOriginal
            global ws
            ws = wb.active

            # PERMITO LA ESCRITURA AL WIDGET DE IMPRESION
            printData.config(state=NORMAL)

            # MEJORO EL MANEJO DE LA OBTENCION DE DATA DEL WIDGET
            numberCodigoEntryVar = numberCodigoEntry.get("1.0", "end-1c")

            codigoFinal = literalCodigoEntry.get("1.0", "end-1c").upper() + "." + str(
                numberCodigoEntryVar)  # CODIGO DE PLANTILLA EXCEL
            literalCodigoEntry.tag_add("tag_name", "1.0", "end")
            print("COMIENZO DESDE EL CODIGO: ", codigoFinal)

            # CHEQUEO QUE LOS WIDGETS DE MARCA Y PRECIO TENGAN INPUT
            if priceEntry2.get() == "":
                tkinter.messagebox.showinfo(message="FALTA AGREGAR, EL PRECIO, ARTICULO NO AGREGADO AL EXCEL!",
                                            title="HEY!!!")
                # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                printData.config(state=DISABLED)
                break

            if typePickedOriginal == "":
                tkinter.messagebox.showinfo(
                    message="FALTA POR AGREGAR, EL TIPO, REEINGRESE DATOS, ARTICULO NO AGREGADO AL EXCEL!",
                    title="HEY!!!")
                # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                printData.config(state=DISABLED)
                break

            cordenada_Excel = wordfinder(codigoFinal)  # CAMBIO A COORDENADAS CON FUNCION WORDFINDER

            # CAMBIO DE TUPLA A CELDA EXCEL

            celdaCompletaCodigo = cambioABC(cordenada_Excel[0]) + str(cordenada_Excel[1])
            celdaCompletaPrenda = cambioABC(cordenada_Excel[0] + 2) + str(cordenada_Excel[1])
            celdaCompletaPrecio = cambioABC(cordenada_Excel[0] + 8) + str(cordenada_Excel[1])
            celdaCompletaPorcentaje = cambioABC(cordenada_Excel[0] + 9) + str(cordenada_Excel[1])
            print(celdaCompletaPrenda)

            # CHEQUEO SI LA CELDA TIENE VALOR DE PRENDA
            if ws[celdaCompletaPrenda].value != None:
                error1 = "ESTA ES UNA PRENDA YA INGRESADA DESEA SOBREESCRIBIR? CODIGO DE VENTA: " + str(
                    val(cordenada_Excel[1], cordenada_Excel[0]))
                r = tkinter.messagebox.askquestion(message=error1, title="HEY!!!")
                printData.insert(END,
                                 "------------------------------------------------ERROR------------------------"
                                 "-------------------------------------------")
                printData.insert(END, "\n")
                printData1 = "MARCA DEL ARTICULO---PRECIO DE LA FACTURA---NOMBRE DE LA PRENDA --- PRECIO:"
                printData.insert(END, printData1)
                printData.insert(END, "\n")
                printData2 = str(val(cordenada_Excel[1], cordenada_Excel[0] + 1)) + "----------------------" + \
                             str(val(cordenada_Excel[1], cordenada_Excel[0] + 3)) + "$  ----------------------" + \
                             str(val(cordenada_Excel[1], cordenada_Excel[0] + 2)) + "----------------------" + \
                             str(val(cordenada_Excel[1], cordenada_Excel[0] + 8)) + "$ "
                printData.insert(END, printData2)
                printData.insert(END, "\n")
                printData.insert(END,
                                 "------------------------------------------------ERROR---------------------------"
                                 "----------------------------------------")
                printData.yview(END)

                if r == 'no':
                    printData.insert(END, "\n")
                    printData.insert(END, "\n")
                    printData.insert(END,
                                     "------------------------------------------------ERROR----------------------"
                                     "---------------------------------------------")
                    printData.insert(END, "\n")
                    printData.insert(END, "NO SE MODIFICO EL EXCEL")
                    printData.insert(END, "\n")
                    printData.insert(END,
                                     "------------------------------------------------ERROR----------------------"
                                     "---------------------------------------------")
                    printData.insert(END, "\n")
                    printData.yview(END)

                    # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                    printData.config(state=DISABLED)

                    break

            # FINAL ITEMS
            finalItems()
            global finalItem
            print(finalItem)

            # Ingreso de la prenda por el usuario
            usr_prenda = finalItem

            # ingero precio usuario
            usr_precio = float(priceEntry2.get())

            if (val(cordenada_Excel[1], cordenada_Excel[0] + 7) == None):
                error2 = "ESTA FACTURA NO HA SIDO INGRESADA AUN, CHEQUEAR EL CODIGO: " + str(
                    val(cordenada_Excel[1], cordenada_Excel[0]))
                tkinter.messagebox.showinfo(message=error2, title="HEY!!!")

                break

            # porcentaje

            porcentaje = (usr_precio * 100) / val(cordenada_Excel[1], cordenada_Excel[0] + 7) - 100
            ws[celdaCompletaPorcentaje] = porcentaje

            # chequeo de porcentaje menor a 100%

            if (porcentaje < 100):
                printData9 = ("LA GANANCIA ES BAJA, MENOR AL 100%: " + str(porcentaje) + "%, DESEA CONTINUAR?")
                r = tkinter.messagebox.askquestion(message=printData9,
                                                   title="HEY!!!")
                if r == 'no':
                    printData.insert(END, "\n")
                    printData.insert(END,
                                     "------------------------------------------------ERROR----------"
                                     "---------------------------------------------------------")
                    printData.insert(END, "\n")
                    printData.insert(END, "NO SE MODIFICO EL EXCEL")
                    printData.insert(END, "\n")
                    printData.insert(END,
                                     "------------------------------------------------ERROR---------"
                                     "----------------------------------------------------------")
                    printData.insert(END, "\n")
                    printData.yview(END)
                    # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
                    printData.config(state=DISABLED)
                    break

            # Agrego datos a la tabla excel
            ws[celdaCompletaPrenda] = usr_prenda.upper()
            ws[celdaCompletaPrecio] = usr_precio
            ws[celdaCompletaPorcentaje] = porcentaje

            print()
            print()
            printData.insert(END, "\n")
            printData.insert(END, "\n")
            printData.insert(END,
                             "-------------------------------------------------------------------------"
                             "---------------------------------------------------")
            prinData5 = ("SE AGREGO CORRECTAMENTE: " + str(val(cordenada_Excel[1], cordenada_Excel[0] + 2)))
            printData.insert(END, prinData5)
            printData.insert(END, "\n")
            printData6 = ("MARCA: " + str(val(cordenada_Excel[1], cordenada_Excel[0] + 1)))
            printData.insert(END, printData6)
            printData.insert(END, "\n")
            printData7 = ("EL PRECIO ESTABLECIDO ES: " + str(val(cordenada_Excel[1], cordenada_Excel[0] + 8)) + "$")
            printData.insert(END, printData7)
            printData.insert(END, "\n")
            printData.insert(END,
                             "--------------------------------------------------------------------------------"
                             "--------------------------------------------")
            printData.insert(END, "\n")
            printData8 = ("CODIGO DE VENTA: " + str(val(cordenada_Excel[1], cordenada_Excel[0])))
            printData.insert(END, printData8)
            printData.insert(END, "\n")
            printData.insert(END,
                             "---------------------------------------------------------------------------"
                             "-------------------------------------------------")
            printData.yview(END)

            # GUARDO EL ARCHIVO CON EL PRODUCTO NUEVO AGREGADO
            wb.save(filename.replace(" ", ""))
            ws = wb.close

            # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION
            printData.config(state=DISABLED)

            break

    addItemTypeButton = Button(rootMain, text=" ADD ITEM ", command=addItem)
    addItemTypeButton.place(x=910,y = 340)
    addItemTypeButton.config(highlightbackground="#FDE5E6")

    rootMain.mainloop()


# ----------------------------------------------------------------------------------------------------------------------

def ventanaSearch():
    deleteRoot = Toplevel()
    deleteRoot.title('CRIS BPUTIQUE DELETE INVENTORY')
    deleteRoot.geometry('500x500')
    deleteRoot.resizable(0, 0)
    print("VENTANA CLIENTES CREADA")

    if filename == "":
        print("error filename")
        r = tkinter.messagebox.askquestion(
            message="ARCHIVO DE EXCEL NO AGREGADO!!! NO SE ABRIRA LA OPCION, AGREGAR AHORA?", title="HEY!!!")
        if (r == 'yes'):
            inputExcel()
        if (r == 'no'):
            deleteRoot.destroy()


    img = PhotoImage(file='FONDOSEARCH.png')
    fondo = Label(deleteRoot, image=img)
    fondo.place(x=0, y=0)

    # BUSQUEDA DEL CODIGO EN EXCEL
    def wordfinder(searchString):
        global wb
        global ws
        ws = wb.active
        while (True):
            buscarColumn = 0
            buscarFila = 0
            count = 0

            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    if searchString == ws.cell(i, j).value:
                        print("SE ENCONTRO EL CODIGO EN LA HOJA EXCEL")
                        print("/////////////////////////////////////////")
                        # print(ws.cell(i,j))
                        buscarColumn = i
                        buscarFila = j
                        count += 1
                        # print("Columna ", buscarColumn)
                        # print("Fila", buscarFila)

            if buscarFila == 0 and buscarFila == 0:
                print("NO SE ENCONTRARON MAS CODIGOS EN LA HOJA DE EXCEL")
                error4 = "ERROR, NO SE ENCONTRO EL CODIGO"
                tkinter.messagebox.showinfo(message=error4, title="HEY!!!")
                buscarColumn = 0
                buscarFila = 0
                print("/////////////////////////////////////////")

                break



            return buscarFila, buscarColumn, count

    # defino una nueva funcion que cambia numeros por letras
    def cambioABC(cambio):
        if cambio == 1:
            cambioNumaLetra = "A"
        elif cambio == 2:
            cambioNumaLetra = "B"
        elif cambio == 3:
            cambioNumaLetra = "C"
        elif cambio == 4:
            cambioNumaLetra = "D"
        elif cambio == 5:
            cambioNumaLetra = "E"
        elif cambio == 6:
            cambioNumaLetra = "F"
        elif cambio == 7:
            cambioNumaLetra = "G"
        elif cambio == 8:
            cambioNumaLetra = "H"
        elif cambio == 9:
            cambioNumaLetra = "I"
        elif cambio == 10:
            cambioNumaLetra = "J"
        elif cambio == 11:
            cambioNumaLetra = "K"
        elif cambio == 12:
            cambioNumaLetra = "L"
        elif cambio == 13:
            cambioNumaLetra = "M"
        elif cambio == 14:
            cambioNumaLetra = "N"
        elif cambio == 15:
            cambioNumaLetra = "O"
        elif cambio == 16:
            cambioNumaLetra = "P"
        elif cambio == 17:
            cambioNumaLetra = "Q"
        elif cambio == 18:
            cambioNumaLetra = "R"
        elif cambio == 19:
            cambioNumaLetra = "S"
        elif cambio == 20:
            cambioNumaLetra = "T"
        elif cambio == 21:
            cambioNumaLetra = "U"
        elif cambio == 22:
            cambioNumaLetra = "V"
        elif cambio == 23:
            cambioNumaLetra = "W"
        elif cambio == 24:
            cambioNumaLetra = "X"
        elif cambio == 25:
            cambioNumaLetra = "Y"
        elif cambio == 26:
            cambioNumaLetra = "Z"
        else:
            cambioNumaLetra = "A"

        return cambioNumaLetra

    # creo un color a rellenar la celda
    yellowFill = PatternFill(start_color='00FFFF00',
                             end_color='00FFFF00',
                             fill_type='solid')
    whiteFill = PatternFill(start_color='00FFFFFF',
                            end_color='00FFFFFF',
                            fill_type='solid')

    # creo una funcion que retorna el valor de una celda con numeros
    def val(x, y):
        return ws.cell(row=x, column=y).value


    def pintarCodigo():
        ws = wb.active
        printData.config(state=NORMAL)
        printData2.config(state=NORMAL)
        printData3.config(state=NORMAL)
        codigoUsr = inputCodigo.get()
        print("CODIGO A BUSCAR Y PINTAR " + codigoUsr)


        word = wordfinder(codigoUsr.upper())
        wordMas = word[0]

        for i in range(0, 10):
            ws = wb.active

            if word[0] != 0:
                if i == 0:
                    palCompleta = cambioABC(word[0]) + str(word[1])
                    if ws[palCompleta].fill == yellowFill:
                        error1 = "ESTE ARTICULO YA ESTA PINTADO"
                        tkinter.messagebox.showinfo(message=error1, title="HEY!!!")
                        print("ESTE ARTICULO YA ESTA PINTADO")
                        break

                    articulo = ws.cell(word[1], 3).value
                    marca = ws.cell(word[1], 2).value
                    precio = val(word[1], 9)

                    print()
                    print("EL ARTICULO PINTADO ES: ", articulo, "MARCA: ", marca)
                    print("EL PRECIO DEL ARTICULO ES : ", precio, "$")
                    print()
                    print("/////////////////////////////////////////")


                    printIt1 = "EL ARTICULO PINTADO ES: " + articulo + " MARCA: "+ marca
                    printData.insert(END, "\n")
                    printData.insert(END, printIt1)
                    printData.yview(END)

                    printIt2 = "PRECIO ES : " + str(precio) + "$"
                    printData2.insert(END, "\n")
                    printData2.insert(END, printIt2)
                    printData2.yview(END)

                    printIt3 = "CELDA ES  : " + str(palCompleta)
                    printData3.insert(END, "\n")
                    printData3.insert(END, printIt3)
                    printData3.yview(END)

                    # asigno la celda del usuraio y la pinto de color
                    ws[palCompleta].fill = yellowFill
                    print("RECUERDE CHEQUEAR EL ARCHIVO, SE ENCONTRO EL CODIGO EN LA CELDA: ", palCompleta)
                    print("--------------------------------------------------------------")


                else:
                    wordMas += 1

                    palCompleta = cambioABC(wordMas) + str(word[1])
                    # print("Esta es la celda completa", palCompleta)

                # asigno la celda del usuraio y la pinto de color
                ws[palCompleta].fill = yellowFill



            else:
                error2 = "ERROR NO SE ENCONTRO EL CODIGO"
                tkinter.messagebox.showinfo(message=error2, title="HEY!!!")
                print("Error, no se encontro el codigo")
                print("--------------------------------------------------------------")
            wb.save(filename.replace(" ", ""))  # guardo el excel
            ws = wb.close
            printData.config(state=DISABLED)
            printData2.config(state=DISABLED)
            printData3.config(state=DISABLED)

    def despintarCodigo():
        # PERMITO LA ESCRITURA AL WIDGET DE IMPRESION
        global ws
        ws = wb.active
        printData.config(state=NORMAL)
        printData2.config(state=NORMAL)
        printData3.config(state=NORMAL)
        print()
        codigoUsr = inputCodigo.get()
        print("CODIGO A BUSCAR Y PINTAR " + codigoUsr)
        word = wordfinder(codigoUsr.upper())

        # print(word) #retorno una tupla de mi funcion
        # print(type(word))
        # print(word[0], "PALABRA EN 0")

        # creo variable palCompleta, me da la coordenada exacta de la busqueda del usuario pasada a excel
        # #el loop pinta la celda y las que le siguen
        wordMas = word[0]

        # mismo programa pero borro
        for i in range(0, 10):
            if word[0] != 0:
                if i == 0:
                    palCompleta = cambioABC(word[0]) + str(word[1])
                    if ws[palCompleta].fill == whiteFill:
                        error3 = "ESTE ARTICULO YA ESTA ELIMINADO"
                        tkinter.messagebox.showinfo(message=error3, title="HEY!!!")
                        print("ESTE ARTICULO YA ESTA ELIMINADO")
                        break

                    articulo = ws.cell(word[1], 3).value
                    marca = ws.cell(word[1], 2).value
                    precio = val(word[1], 9)

                    print()
                    print("EL ARTICULO ELIMINADO ES: ", articulo, "MARCA: ", marca)
                    print("EL PRECIO DEL ARTICULO ES : ", precio, "$")
                    print()
                    print("/////////////////////////////////////////")

                    printIt1 = "EL ARTICULO ElIMINADO ES: " + articulo + " MARCA: " + marca
                    printData.insert(END, "\n")
                    printData.insert(END, printIt1)
                    printData.yview(END)

                    printIt2 = "PRECIO ES : " + str(precio) + "$"
                    printData2.insert(END, "\n")
                    printData2.insert(END, printIt2)
                    printData2.yview(END)

                    printIt3 = "CELDA ES  : " + str(palCompleta)
                    printData3.insert(END, "\n")
                    printData3.insert(END, printIt3)
                    printData3.yview(END)

                    # asigno la celda del usuraio y la pinto de color
                    ws[palCompleta].fill = yellowFill
                    print("RECUERDE CHEQUEAR EL ARCHIVO, SE ENCONTRO EL CODIGO EN LA CELDA: ", palCompleta)
                    print("--------------------------------------------------------------")


                else:
                    wordMas += 1

                    palCompleta = cambioABC(wordMas) + str(word[1])
                    # print("Esta es la celda completa", palCompleta)

                    # asigno la celda del usuraio y la pinto de color
                ws[palCompleta].fill = whiteFill





            else:
                error4 = "ERROR, NO SE ENCONTRO EL CODIGO"
                tkinter.messagebox.showinfo(message=error4, title="HEY!!!")
                print("Error, no se encontro el codigo")
                print("--------------------------------------------------------------")
                break

            wb.save(filename.replace(" ", ""))  # guardo el excel
            printData.config(state=DISABLED)
            printData2.config(state=DISABLED)
            printData3.config(state=DISABLED)

        ws = wb.close





    # CODIGO INPUT
    inputCodigo = ttk.Entry(deleteRoot, width=5, background="#FFB1BE")
    inputCodigo.place(x=100, y=85)
    # BOTON PINTAR
    pintarButton = Button(deleteRoot, text="BAJAR", command=pintarCodigo)
    pintarButton.place(x=248, y=85)
    pintarButton.config(highlightbackground="#FDE5E6")

    # BOTON DESPINTAR\
    despintarButton = Button(deleteRoot, text="ELIMINAR", command=despintarCodigo)
    despintarButton.place(x=345, y=85)
    despintarButton.config(highlightbackground="#FDE5E6")

    #FONT
    fontTupleEntry = ("Gotham Medium", 16)

    # DISPLAY RECUADRO DE IMPRESION DE DATA ARTICULO
    printData = Text(deleteRoot, height=2, width=35, font=fontTupleEntry, wrap=WORD, background="#FDE5E6")
    printData.place(x=70, y=185)
    printData.config(highlightbackground="#FDE5E6")
    printData.config(state=DISABLED)  # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION

    # DISPLAY RECUADRO DE IMPRESION DE PRECIO
    printData2 = Text(deleteRoot, height=2, width=10, font=fontTupleEntry, wrap=WORD, background="#FDE5E6")
    printData2.place(x=75, y=276)
    printData2.config(highlightbackground="#FDE5E6")
    printData2.config(state=DISABLED)  # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION

    # DISPLAY RECUADRO DE IMPRESION DE CELDA
    printData3 = Text(deleteRoot, height=2, width=10, font=fontTupleEntry, wrap=WORD, background="#FDE5E6")
    printData3.place(x=75, y=365)
    printData3.config(highlightbackground="#FDE5E6")
    printData3.config(state=DISABLED)  # DENEGO LA ESCRITURA AL WIDGET DE IMPRESION

    # ABRIR EXCEL
    def openFile():
        subprocess.run(['open', filename], check=True)

    openExcel = Button(deleteRoot, text="ABRIR EXCEL", command=openFile)
    openExcel.place(x=345, y=460)
    openExcel.config(highlightbackground="#FDE5E6")

    deleteRoot.mainloop()




# -----------------------------------------------------------------------------------------------------

def clients():
    clientsRoot = Toplevel()
    clientsRoot.title('CRIS BPUTIQUE CLIENTS')
    clientsRoot.geometry('1366x768')
    clientsRoot.resizable(0, 0)
    print("VENTANA CLIENTES CREADA")

    if filename == "":
        print("error filename")
        r = tkinter.messagebox.askquestion(
            message="ARCHIVO DE EXCEL NO AGREGADO!!! NO SE ABRIRA LA OPCION, AGREGAR AHORA?", title="HEY!!!")
        if (r == 'yes'):
            inputExcel()
        if (r == 'no'):
            clientsRoot.destroy()

    # abro el excel
    global ws
    global wb
    ws = wb.active

    # BUSQUEDA DEL CODIGO EN EXCEL
    def wordfinder(searchString):
        global wb
        global ws
        ws = wb.active
        while (True):
            buscarColumn = 0
            buscarFila = 0
            count = 0

            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    if searchString == ws.cell(i, j).value:
                        print("SE ENCONTRO EL CODIGO EN LA HOJA EXCEL")
                        print("/////////////////////////////////////////")
                        # print(ws.cell(i,j))
                        buscarColumn = i
                        buscarFila = j
                        count += 1
                        # print("Columna ", buscarColumn)
                        # print("Fila", buscarFila)

            if buscarFila == 0 and buscarFila == 0:
                print("NO SE ENCONTRARON MAS CODIGOS EN LA HOJA DE EXCEL")
                buscarColumn = 0
                buscarFila = 0
                print("/////////////////////////////////////////")

                break

            return buscarFila, buscarColumn, count

    def cambioABC(cambio):
        if cambio == 1:
            cambioNumaLetra = "A"
        elif cambio == 2:
            cambioNumaLetra = "B"
        elif cambio == 3:
            cambioNumaLetra = "C"
        elif cambio == 4:
            cambioNumaLetra = "D"
        elif cambio == 5:
            cambioNumaLetra = "E"
        elif cambio == 6:
            cambioNumaLetra = "F"
        elif cambio == 7:
            cambioNumaLetra = "G"
        elif cambio == 8:
            cambioNumaLetra = "H"
        elif cambio == 9:
            cambioNumaLetra = "I"
        elif cambio == 10:
            cambioNumaLetra = "J"
        elif cambio == 11:
            cambioNumaLetra = "K"
        elif cambio == 12:
            cambioNumaLetra = "L"
        elif cambio == 13:
            cambioNumaLetra = "M"
        elif cambio == 14:
            cambioNumaLetra = "N"
        elif cambio == 15:
            cambioNumaLetra = "O"
        elif cambio == 16:
            cambioNumaLetra = "P"
        elif cambio == 17:
            cambioNumaLetra = "Q"
        elif cambio == 18:
            cambioNumaLetra = "R"
        elif cambio == 19:
            cambioNumaLetra = "S"
        elif cambio == 20:
            cambioNumaLetra = "T"
        elif cambio == 21:
            cambioNumaLetra = "U"
        elif cambio == 22:
            cambioNumaLetra = "V"
        elif cambio == 23:
            cambioNumaLetra = "W"
        elif cambio == 24:
            cambioNumaLetra = "X"
        elif cambio == 25:
            cambioNumaLetra = "Y"
        elif cambio == 26:
            cambioNumaLetra = "Z"
        else:
            cambioNumaLetra = "A"

        return cambioNumaLetra

    # RETORNO DE CELDA CON COORDENADAS
    def val(x, y):
        return ws.cell(row=x, column=y).value

    def pruebaRueleta():
        # print(ruletaDeUsuarios.get())
        busqueda = wordfinder("A.1")
        countadorNombres = 0
        listaUsuarios = []

        fontTupleEntry2 = ("Gotham Medium", 60)
        printData = Text(clientsRoot, height=3, width=34, wrap=WORD, background="#FFB1BE")
        printData.place(x=60, y=92)
        printData.config(highlightbackground="#FFB1BE", font=fontTupleEntry2)

        while (True):
            printData.config(state=NORMAL)
            nombresCeldas = val((busqueda[1] + countadorNombres), busqueda[0])
            nombresUsuariosFinales = val((busqueda[1] + countadorNombres), (busqueda[0] + 1))
            facturaFinal = val((busqueda[1] + countadorNombres), (busqueda[0] + 2))
            abonoFinal = val((busqueda[1] + countadorNombres), (busqueda[0] + 3))
            descuentoFinal = val((busqueda[1] + countadorNombres), (busqueda[0] + 4))
            aPagarFinal = val((busqueda[1] + countadorNombres), (busqueda[0] + 5))
            if str(facturaFinal) == "None":
                facturaFinal = ""
            if str(abonoFinal) == "None":
                abonoFinal = ""
            if str(descuentoFinal) == "None":
                descuentoFinal = ""
            if str(aPagarFinal) == "None":
                aPagarFinal = ""
            printdata1 = str(nombresUsuariosFinales) + "   " + str(facturaFinal) + "   " + str(
                abonoFinal) + "   " + str(descuentoFinal) + "   " + str(aPagarFinal)
            printData.insert(END, printdata1)
            printData.insert(END, "\n")
            printData.config(state=DISABLED)

            print(str(nombresCeldas) + " CELDA")
            print(nombresUsuariosFinales)

            countadorNombres += 1

            nombresUsuariosFinales = str(nombresUsuariosFinales)
            listaUsuarios.append(nombresUsuariosFinales)

            if (str(nombresUsuariosFinales) == "None"):
                listaUsuarios.remove('None')
                print(listaUsuarios)
                return listaUsuarios, countadorNombres

    def obtenerAbono():

        while (True):
            global ws
            global wb
            ws = wb.active

            celda = wordfinder(str(ruletaDeUsuarios.get()))
            print(celda)
            if str(celda) == "None":
                tkinter.messagebox.showinfo(message="NO SE HA ELEGIDO NINGUN USUARIO", title="HEY!!!")
                break

            else:
                abono = inputFactura.get()
                if (abono == ""):
                    tkinter.messagebox.showinfo(message="NO SE AGREGO NINGUN ABONO", title="HEY!!!")
                    break

                else:
                    abonoCeldas = cambioABC(celda[0] + 2) + str(celda[1])
                    porPagarCeldas = cambioABC(celda[0] + 4) + str(celda[1])
                    abonoAnterior = (ws.cell((celda[1]), celda[0] + 2).value)
                    historialDeAbonosAnterior = (ws.cell((celda[1]), celda[0] + 6).value)
                    abonoFinal = float(abono) + float(abonoAnterior)
                    factura = (ws.cell((celda[1]), celda[0] + 1).value)
                    porPagarAnterior = (ws.cell((celda[1]), celda[0] + 4).value)
                    porPagar = float(porPagarAnterior) - float(abono)

                    # HISTORIAL DE ABONOS
                    historialDeAbonosCelda = cambioABC(celda[0] + 6) + str(celda[1])
                    historialDeAbonos = str(historialDeAbonosAnterior) + "+" + str(abono)
                    if float(abono) > float(porPagarAnterior):
                        tkinter.messagebox.showinfo(
                            message="EL ABONO INGRESADO ES MAYOR AL PRECIO POR PAGAR, REVISAR EL ABONO INGRESADO",
                            title="HEY!!!")
                        break

                    print(abonoFinal)
                    ws[abonoCeldas] = abonoFinal
                    ws[porPagarCeldas] = porPagar

                    # AGREGO HISTORIAL DE ABONOS
                    ws[historialDeAbonosCelda] = historialDeAbonos

                    ws = wb.save(filename.replace(" ", ""))
                    ws = wb.close()
                    mensaje = "SE AGREGO EL ABONO DE: $" + str(abono) + " AL CLIENTE: " + \
                              str(ruletaDeUsuarios.get()) + \
                              ". LA SUMA DE ABONOS PARA ESTE CLIENTE ES DE: $" + str(abonoFinal)
                    tkinter.messagebox.showinfo(message=mensaje, title="HEY!!!")

                    if float(porPagar) == 0:
                        tkinter.messagebox.showinfo(message="EL CLIENTE HA PAGADO SU DEUDA!!!", title="HEY!!!")

                    pruebaRueleta()

                    break

    def nuevoCliente():
        while (True):
            global ws
            global wb
            ws = wb.active
            nombre = inputCliente.get()
            if nombre == "":
                tkinter.messagebox.showinfo(message="NO SE AGREGO NINGUN NOMBRE", title="HEY!!!")
                break
            print(nombre)

            precioAPagar = inputPrecioAPagar.get()
            if precioAPagar == "":
                tkinter.messagebox.showinfo(message="NO SE AGREGO NINGUN PRECIO A PAGAR", title="HEY!!!")
                break
            print(precioAPagar)

            descuento = inputDescuento.get()
            if descuento == "":
                descuento = "0"
            print(descuento)

            print(pruebaRueleta()[1], "IMPRIMO")

            codigoCeldaAAgrefar = "A." + str(pruebaRueleta()[1])
            print(codigoCeldaAAgrefar)

            coordenadasCeldaEncontrada = wordfinder(codigoCeldaAAgrefar)
            print(coordenadasCeldaEncontrada)

            nombreCelda = cambioABC(coordenadasCeldaEncontrada[0] + 1) + str(coordenadasCeldaEncontrada[1])
            precioAPagarCelda = cambioABC(coordenadasCeldaEncontrada[0] + 2) + str(coordenadasCeldaEncontrada[1])
            abonoCelda = cambioABC(coordenadasCeldaEncontrada[0] + 3) + str(coordenadasCeldaEncontrada[1])
            descuentoCelda = cambioABC(coordenadasCeldaEncontrada[0] + 4) + str(coordenadasCeldaEncontrada[1])
            porPagarCelda = cambioABC(coordenadasCeldaEncontrada[0] + 5) + str(coordenadasCeldaEncontrada[1])
            historialDeAbonosCelda = cambioABC(coordenadasCeldaEncontrada[0] + 7) + str(coordenadasCeldaEncontrada[1])

            ws[nombreCelda] = nombre
            descuentoFinal = float(precioAPagar) - float(descuento)
            ws[precioAPagarCelda] = str(precioAPagar)
            ws[abonoCelda] = "0"
            ws[descuentoCelda] = descuento
            ws[porPagarCelda] = str(descuentoFinal)
            ws[historialDeAbonosCelda] = "+0"

            ws = wb.save(filename.replace(" ", ""))
            ws = wb.close()
            pruebaRueleta()
            tkinter.messagebox.showinfo(message="NUEVO CLIENTE AGREGADO", title="HEY!!!")
            ruletaDeUsuarios['values'] = pruebaRueleta()[0]

            break

    img = PhotoImage(file='FONDOCLIENTES.png')
    fondo = Label(clientsRoot, image=img)
    fondo.place(x=0, y=0)

    # Ruleta
    style = ttk.Style()
    style.configure("TCombobox", fieldbackground="#FFB1BE", background="#FFB1BE")
    global listaUsuarios
    countryvar = StringVar()
    ruletaDeUsuarios = ttk.Combobox(clientsRoot, width=53, textvariable=countryvar, background="#FFB1BE")

    ruletaDeUsuarios['values'] = pruebaRueleta()[0]
    ruletaDeUsuarios.place(x=70, y=710)
    # FACTURA INPUT
    inputFactura = ttk.Entry(clientsRoot, width=5, background="#FFB1BE")
    inputFactura.place(x=705, y=710)
    # BOTON AGREGAR ABONO
    agregarAbonoButton = Button(clientsRoot, text="AGREGAR ABONO", command=obtenerAbono)
    agregarAbonoButton.place(x=965, y=710)
    agregarAbonoButton.config(highlightbackground="#FFB1BE")

    # CLIENTE ENTRY
    inputCliente = ttk.Entry(clientsRoot, width=20, background="#FFB1BE")
    inputCliente.place(x=180, y=510)
    # PRECIO A PAGAR ENTRY
    inputPrecioAPagar = ttk.Entry(clientsRoot, width=5, background="#FFB1BE")
    inputPrecioAPagar.place(x=660, y=510)
    # DESCUENTO ENTRY
    inputDescuento = ttk.Entry(clientsRoot, width=5, background="#FFB1BE")
    inputDescuento.place(x=1060, y=510)
    # BOTON AGREGAR CLIENTE
    agregarClienteButton = Button(clientsRoot, text="AGREGAR VENTA", command=nuevoCliente)
    agregarClienteButton.place(x=1155, y=582)
    agregarClienteButton.config(highlightbackground="#FFB1BE")

    # ABRIR EXCEL
    def openFile():
        subprocess.run(['open', filename], check=True)

    openExcel = Button(clientsRoot, text="ABRIR EXCEL", command=openFile)
    openExcel.place(x=1180, y=710)
    openExcel.config(highlightbackground="#FFB1BE")

    # guardado y cerrado de excel
    ws = wb.save(filename.replace(" ", ""))
    ws = wb.close
    clientsRoot.mainloop()


# ----------------------------------------------------------------------------------------------------------------------

# MENU PRINCIPAL
def menuPrincipal():
    global root
    root = Tk()
    root.title('CRIS BOUTIQUE')
    root.config(background='#FDE5E6')
    root.geometry("510x310")
    root.resizable(0, 0)
    root.eval('tk::PlaceWindow . center')

    # MENU ILLUSTRATOR
    img = PhotoImage(file='FONDO-MENU.png')
    fondo = Label(root, image=img)
    fondo.config(background='#FFB1BE')
    fondo.place(x=0, y=0)

    # MENU PRINCIPAL BOTONES
    searchButtonWindow = Button(root, text="DELETE INVENTORY", command=ventanaSearch)
    searchButtonWindow.place(x=300, y=220)
    searchButtonWindow.config(highlightbackground="#FFB1BE")

    addClothesAndBillsWindow = Button(root, text="ADD BILLS OR CLOTHES TO INDEX", command=addClothesAndBills)
    addClothesAndBillsWindow.place(x=50, y=220)
    addClothesAndBillsWindow.config(highlightbackground="#FFB1BE")

    searchExcelFile = Button(root, text="OPEN EXCEL", command=inputExcel)
    searchExcelFile.place(x=200, y=120)
    searchExcelFile.config(highlightbackground="#FFB1BE")

    addClients = Button(root, text="CLIENTS", command=clients)
    addClients.place(x=350, y=260)
    addClients.config(highlightbackground="#FFB1BE")

    root.mainloop()


# LLAMADA A LA FUNCION MENU PRINCIPAL
menuPrincipal()


